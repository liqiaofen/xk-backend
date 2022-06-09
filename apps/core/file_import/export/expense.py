from abc import ABC, abstractmethod

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class BaseExport(ABC):

    def __init__(self, queryset, model=None, **kwargs):
        self.queryset = queryset
        self.model = model
        self.kwargs = kwargs
        self.workbook = Workbook()

    @staticmethod
    def set_header(sheet, columns, row_num=1):
        for col_num, (title, width) in enumerate(columns, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = title
            # cell.font =
            # 设置列宽
            column_letter = get_column_letter(col_num)
            column_dimensions = sheet.column_dimensions[column_letter]
            column_dimensions.width = width

    @staticmethod
    def set_row(sheet, data: list, row_num):
        for col_num, (value, font) in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = font
            # cell.alignment = self.style["wrapped_alignment"]

    def create_sheet(self, title='Sheet', columns=None, row_num=1):
        sheet = self.workbook[title]
        sheet.title = title  # 将缺省Sheet改名
        self.set_header(sheet, columns, row_num)

    @abstractmethod
    def deal_data(self):
        pass


class ExpenseExport(BaseExport):

    def __init__(self, queryset, model=None, **kwargs):
        super(ExpenseExport, self).__init__(queryset, model=model, **kwargs)

    def deal_data(self):
        columns = [
            # ('订单号', 30),
        ]
        row_num = 1
        sheet = self.create_sheet(columns=columns, row_num=row_num)

        for obj in self.queryset:
            row_num += 1

            # 每行按顺序显示的值和样式
            row = [
                (obj['number'], 'Normal'),
            ]
            # 渲染订单列表数据
            self.set_row(sheet, row, row_num)

        sheet.freeze_panes = sheet['A2']
